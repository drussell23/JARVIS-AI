#!/usr/bin/env python3
"""
Export Voice Unlock Metrics to Excel
=====================================
Converts all JSON metrics logs to a comprehensive Excel workbook with multiple sheets.

Usage:
    python export_metrics_to_excel.py                    # Export today's data
    python export_metrics_to_excel.py --date 2025-11-13  # Export specific date
    python export_metrics_to_excel.py --all              # Export all available dates
"""

import json
import sys
from pathlib import Path
from datetime import datetime, timedelta
from typing import List, Dict, Any
import argparse

try:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ Missing required packages. Installing...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "pandas", "openpyxl"])
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter


class MetricsToExcelExporter:
    """Export voice unlock metrics from JSON to formatted Excel workbook"""

    def __init__(self, log_dir: str = None):
        """Initialize exporter with log directory"""
        if log_dir is None:
            self.log_dir = Path.home() / ".jarvis/logs/unlock_metrics"
        else:
            self.log_dir = Path(log_dir)

        # Output directory for Excel files
        self.output_dir = self.log_dir / "excel_exports"
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def export_to_excel(self, date: str = None, export_all: bool = False) -> str:
        """
        Export metrics to Excel workbook.

        Args:
            date: Specific date to export (YYYY-MM-DD), defaults to today
            export_all: Export all available dates

        Returns:
            Path to created Excel file
        """
        if export_all:
            return self._export_all_dates()

        if date is None:
            date = datetime.now().strftime("%Y-%m-%d")

        print(f"📊 Exporting metrics for {date}...")

        # Load JSON files
        daily_log = self._load_daily_log(date)
        stats = self._load_stats()
        trends = self._load_trends()

        if not daily_log and not stats and not trends:
            print(f"❌ No data found for {date}")
            return None

        # Create Excel filename
        excel_file = self.output_dir / f"unlock_metrics_{date}.xlsx"

        # Create DataFrames for each sheet
        sheets = {}

        if daily_log:
            sheets['Unlock Attempts'] = self._create_attempts_sheet(daily_log)
            sheets['Stage Performance'] = self._create_stages_sheet(daily_log)
            sheets['Biometrics Detail'] = self._create_biometrics_sheet(daily_log)

        if stats:
            sheets['Overall Statistics'] = self._create_stats_sheet(stats)

        if trends:
            sheets['Confidence Trends'] = self._create_trends_sheet(trends)

        # Write to Excel with formatting
        self._write_excel(sheets, excel_file)

        print(f"✅ Excel file created: {excel_file}")
        print(f"📊 Sheets created: {', '.join(sheets.keys())}")
        return str(excel_file)

    def _load_daily_log(self, date: str) -> List[Dict]:
        """Load daily log JSON file"""
        log_file = self.log_dir / f"unlock_metrics_{date}.json"
        if not log_file.exists():
            return []

        with open(log_file) as f:
            return json.load(f)

    def _load_stats(self) -> Dict:
        """Load aggregated stats JSON file"""
        stats_file = self.log_dir / "unlock_stats.json"
        if not stats_file.exists():
            return {}

        with open(stats_file) as f:
            return json.load(f)

    def _load_trends(self) -> Dict:
        """Load confidence trends JSON file"""
        trends_file = self.log_dir / "confidence_trends.json"
        if not trends_file.exists():
            return {}

        with open(trends_file) as f:
            return json.load(f)

    def _create_attempts_sheet(self, daily_log: List[Dict]) -> pd.DataFrame:
        """Create main unlock attempts sheet"""
        rows = []

        for entry in daily_log:
            rows.append({
                'Timestamp': entry['timestamp'],
                'Date': entry['date'],
                'Time': entry['time'],
                'Day of Week': entry['day_of_week'],
                'Success': '✅' if entry['success'] else '❌',
                'Speaker Name': entry['speaker_name'],
                'Transcribed Text': entry['transcribed_text'],
                'Confidence Score': entry['biometrics']['speaker_confidence'],
                'STT Confidence': entry['biometrics']['stt_confidence'],
                'Threshold': entry['biometrics']['threshold'],
                'Above Threshold': '✅' if entry['biometrics']['above_threshold'] else '❌',
                'Confidence Margin': entry['biometrics']['confidence_margin'],
                'Margin %': entry['biometrics']['confidence_vs_threshold']['margin_percentage'],
                'Total Duration (ms)': entry['performance']['total_duration_ms'],
                'Total Duration (sec)': round(entry['performance']['total_duration_ms'] / 1000, 2),
                'Slowest Stage': entry['performance'].get('slowest_stage', 'N/A'),
                'Audio Quality': entry['quality_indicators']['audio_quality'],
                'Voice Match Quality': entry['quality_indicators']['voice_match_quality'],
                'Overall Confidence': entry['quality_indicators']['overall_confidence'],
                'Trend Direction': entry['biometrics']['confidence_trends'].get('trend_direction', 'N/A'),
                'Avg Last 10': entry['biometrics']['confidence_trends'].get('avg_last_10', 0),
                'Best Ever': entry['biometrics']['confidence_trends'].get('best_ever', 0),
                'Percentile Rank': entry['biometrics']['confidence_trends'].get('current_rank_percentile', 0),
                'Total Stages': entry['stage_summary']['total_stages'],
                'Successful Stages': entry['stage_summary']['successful_stages'],
                'Failed Stages': entry['stage_summary']['failed_stages'],
                'All Stages Passed': '✅' if entry['stage_summary']['all_stages_passed'] else '❌',
                'Error': entry.get('error', ''),
                'Platform': entry['system_info'].get('platform', 'N/A'),
                'STT Engine': entry['system_info'].get('stt_engine', 'N/A'),
                'Session ID': entry['metadata']['session_id'],
            })

        return pd.DataFrame(rows)

    def _create_stages_sheet(self, daily_log: List[Dict]) -> pd.DataFrame:
        """Create detailed stage performance sheet"""
        rows = []

        for entry in daily_log:
            attempt_time = entry['time']
            for stage in entry['processing_stages']:
                rows.append({
                    'Attempt Time': attempt_time,
                    'Stage Name': stage['stage_name'],
                    'Duration (ms)': stage['duration_ms'],
                    'Duration (sec)': round(stage['duration_ms'] / 1000, 2) if stage['duration_ms'] else 0,
                    '% of Total': round(stage['percentage_of_total'], 1),
                    'Success': '✅' if stage['success'] else '❌',
                    'Algorithm Used': stage.get('algorithm_used', 'N/A'),
                    'Function Name': stage.get('function_name', 'N/A'),
                    'Confidence Score': stage.get('confidence_score', 'N/A'),
                    'Threshold': stage.get('threshold', 'N/A'),
                    'Above Threshold': '✅' if stage.get('above_threshold') else ('❌' if stage.get('above_threshold') is False else 'N/A'),
                    'Input Size (bytes)': stage.get('input_size_bytes', 'N/A'),
                    'Output Size (bytes)': stage.get('output_size_bytes', 'N/A'),
                    'Error': stage.get('error_message', ''),
                })

        return pd.DataFrame(rows)

    def _create_biometrics_sheet(self, daily_log: List[Dict]) -> pd.DataFrame:
        """Create detailed biometrics analysis sheet"""
        rows = []

        for entry in daily_log:
            bio = entry['biometrics']
            trends = bio['confidence_trends']

            rows.append({
                'Timestamp': entry['timestamp'],
                'Speaker Confidence': bio['speaker_confidence'],
                'STT Confidence': bio['stt_confidence'],
                'Threshold': bio['threshold'],
                'Above Threshold': '✅' if bio['above_threshold'] else '❌',
                'Confidence Margin': bio['confidence_margin'],
                'Margin %': bio['confidence_vs_threshold']['margin_percentage'],
                'Avg Last 10': trends.get('avg_last_10', 0),
                'Avg Last 30': trends.get('avg_last_30', 0),
                'Trend Direction': trends.get('trend_direction', 'N/A'),
                'Volatility (Std Dev)': trends.get('volatility', 0),
                'Best Ever': trends.get('best_ever', 0),
                'Worst Ever': trends.get('worst_ever', 0),
                'Percentile Rank': trends.get('current_rank_percentile', 0),
                'Success': '✅' if entry['success'] else '❌',
            })

        return pd.DataFrame(rows)

    def _create_stats_sheet(self, stats: Dict) -> pd.DataFrame:
        """Create overall statistics sheet"""
        rows = []

        # Overall stats
        rows.append({
            'Metric': 'Total Attempts',
            'Value': stats.get('total_attempts', 0),
            'Category': 'Overall'
        })
        rows.append({
            'Metric': 'Successful Attempts',
            'Value': stats.get('successful_attempts', 0),
            'Category': 'Overall'
        })
        rows.append({
            'Metric': 'Failed Attempts',
            'Value': stats.get('failed_attempts', 0),
            'Category': 'Overall'
        })
        success_rate = (stats.get('successful_attempts', 0) / stats.get('total_attempts', 1) * 100) if stats.get('total_attempts', 0) > 0 else 0
        rows.append({
            'Metric': 'Success Rate (%)',
            'Value': round(success_rate, 2),
            'Category': 'Overall'
        })
        rows.append({
            'Metric': 'Last Updated',
            'Value': stats.get('last_updated', 'N/A'),
            'Category': 'Overall'
        })

        # Per-speaker stats
        for speaker, speaker_stats in stats.get('speakers', {}).items():
            rows.append({
                'Metric': f'{speaker} - Total Attempts',
                'Value': speaker_stats.get('total_attempts', 0),
                'Category': 'Per Speaker'
            })
            rows.append({
                'Metric': f'{speaker} - Successful Attempts',
                'Value': speaker_stats.get('successful_attempts', 0),
                'Category': 'Per Speaker'
            })
            rows.append({
                'Metric': f'{speaker} - Avg Confidence',
                'Value': round(speaker_stats.get('avg_confidence', 0), 4),
                'Category': 'Per Speaker'
            })
            rows.append({
                'Metric': f'{speaker} - Best Confidence',
                'Value': round(speaker_stats.get('best_confidence', 0), 4),
                'Category': 'Per Speaker'
            })
            rows.append({
                'Metric': f'{speaker} - Worst Confidence',
                'Value': round(speaker_stats.get('worst_confidence', 0), 4),
                'Category': 'Per Speaker'
            })
            rows.append({
                'Metric': f'{speaker} - Avg Duration (ms)',
                'Value': round(speaker_stats.get('avg_duration_ms', 0), 2),
                'Category': 'Per Speaker'
            })

        return pd.DataFrame(rows)

    def _create_trends_sheet(self, trends: Dict) -> pd.DataFrame:
        """Create confidence trends sheet"""
        rows = []

        for speaker, speaker_trends in trends.items():
            confidence_history = speaker_trends.get('confidence_history', [])
            success_history = speaker_trends.get('success_history', [])
            timestamps = speaker_trends.get('timestamps', [])

            for i, (confidence, success, timestamp) in enumerate(zip(confidence_history, success_history, timestamps)):
                rows.append({
                    'Speaker': speaker,
                    'Attempt Number': i + 1,
                    'Timestamp': timestamp,
                    'Confidence': confidence,
                    'Success': '✅' if success else '❌',
                })

        return pd.DataFrame(rows)

    def _write_excel(self, sheets: Dict[str, pd.DataFrame], excel_file: Path):
        """Write DataFrames to Excel with formatting"""
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            for sheet_name, df in sheets.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Apply formatting
        self._apply_formatting(excel_file)

    def _apply_formatting(self, excel_file: Path):
        """Apply professional formatting to Excel workbook"""
        wb = load_workbook(excel_file)

        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Format header row
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Add borders to all cells
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = border

            # Freeze first row
            ws.freeze_panes = ws['A2']

        wb.save(excel_file)

    def _export_all_dates(self) -> str:
        """Export all available dates to a single Excel file"""
        print("📊 Exporting all available dates...")

        # Find all daily log files
        log_files = sorted(self.log_dir.glob("unlock_metrics_*.json"))
        dates = [f.stem.replace("unlock_metrics_", "") for f in log_files]

        if not dates:
            print("❌ No log files found")
            return None

        # Combine all data
        all_data = []
        for date in dates:
            daily_log = self._load_daily_log(date)
            all_data.extend(daily_log)

        # Create comprehensive workbook
        excel_file = self.output_dir / f"unlock_metrics_ALL_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        sheets = {
            'All Unlock Attempts': self._create_attempts_sheet(all_data),
            'All Stage Performance': self._create_stages_sheet(all_data),
            'All Biometrics': self._create_biometrics_sheet(all_data),
            'Overall Statistics': self._create_stats_sheet(self._load_stats()),
            'Confidence Trends': self._create_trends_sheet(self._load_trends()),
        }

        self._write_excel(sheets, excel_file)

        print(f"✅ Comprehensive Excel file created: {excel_file}")
        print(f"📊 Dates included: {len(dates)}")
        print(f"📊 Total attempts: {len(all_data)}")
        return str(excel_file)


def main():
    """Main entry point"""
    parser = argparse.ArgumentParser(description="Export voice unlock metrics to Excel")
    parser.add_argument('--date', type=str, help='Specific date to export (YYYY-MM-DD)')
    parser.add_argument('--all', action='store_true', help='Export all available dates')
    parser.add_argument('--output-dir', type=str, help='Output directory for Excel files')
    parser.add_argument('--open', action='store_true', help='Open Excel file after creation')

    args = parser.parse_args()

    exporter = MetricsToExcelExporter()

    excel_file = exporter.export_to_excel(date=args.date, export_all=args.all)

    if excel_file and args.open:
        import subprocess
        subprocess.run(['open', excel_file])


if __name__ == '__main__':
    main()
